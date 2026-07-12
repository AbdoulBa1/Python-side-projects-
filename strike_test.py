from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
ws["A1"] = "This text will be struck through"
ws["A1"].font = Font(strike=True)
wb.save("strike_test.xlsx")
