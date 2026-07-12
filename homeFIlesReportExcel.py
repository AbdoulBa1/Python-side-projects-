import openpyxl, os
from pathlib import Path

def get_home_folder_size():
    
    filenames_and_sizes = []

    # Loop over everything in the home folder:
    for filename in os.listdir(Path.home()):
        absolute_file_path = Path.home() / filename

        # Skip folders/directories:
        if absolute_file_path.is_dir():
            continue

        # Get file size:
        try:
            file_size = absolute_file_path.stat().st_size
        except Exception:
            # Skip files with permissions errors:
            continue

        # Record filename and size:
        filenames_and_sizes.append((filename, file_size))

    return filenames_and_sizes
print(get_home_folder_size())
def make_excel_report(filenames_and_sizes):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row, (filename,size) in enumerate(filenames_and_sizes, start=1):
        ws[f'A{row}'] = filename
        ws[f'B{row}'] = size
    wb.save('homeFilesReport.xlsx')
make_excel_report(get_home_folder_size())
